import time
# -*- coding: utf-8 -*-
"""review.py — generate รูปรีวิวสินค้าจาก review_products.xlsx แล้วโพส FB"""

import sys, io, os, base64, requests, time, random, re, sqlite3, struct
from datetime import datetime, timezone, timedelta
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from google import genai
import openpyxl
import numpy as np
from overlay_utils import add_overlay

DB_PATH = os.path.join(os.path.dirname(__file__), "page_bot_memory.db")

def init_db():
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS posts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            platform TEXT NOT NULL,
            raw_text TEXT NOT NULL,
            topic TEXT NOT NULL,
            slot TEXT NOT NULL,
            embedding BLOB,
            timestamp INTEGER NOT NULL
        );
        """)
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[DB] Init failed: {e}")

def get_embedding(text: str, client):
    global API_ENABLED
    if not API_ENABLED or not client:
        return None
    try:
        response = client.models.embed_content(
            model="gemini-embedding-2",
            contents=text
        )
        return response.embeddings[0].values
    except Exception as e:
        err_msg = str(e)
        print(f"[Warning] Embedding generation failed: {err_msg[:120]}")
        if "depleted" in err_msg.lower() or "resource_exhausted" in err_msg.lower():
            print("API credits depleted. Disabling API embedding for this run.")
            API_ENABLED = False
        return None

def serialize_vector(vector: list[float]) -> bytes:
    if not vector:
        return None
    return struct.pack(f"{len(vector)}f", *vector)

def deserialize_vector(blob: bytes) -> list[float]:
    if not blob:
        return None
    num_floats = len(blob) // 4
    return list(struct.unpack(f"{num_floats}f", blob))

def is_duplicate(candidate_text: str, client, threshold: float = 0.82, jaccard_threshold: float = 0.60) -> bool:
    global API_ENABLED
    
    # Try getting embedding for candidate
    v_candidate = None
    if API_ENABLED and client:
        vec = get_embedding(candidate_text, client)
        if vec:
            v_candidate = np.array(vec)

    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("SELECT raw_text, embedding FROM posts ORDER BY timestamp DESC LIMIT 50")
        rows = cursor.fetchall()
        conn.close()
    except Exception as e:
        print(f"[DB] Fetch failed: {e}")
        rows = []
        
    if not rows:
        return False
        
    for raw_text, blob in rows:
        # If vector check is available
        if v_candidate is not None and blob is not None:
            try:
                v_hist = np.array(deserialize_vector(blob))
                dot_product = np.dot(v_candidate, v_hist)
                norm_c = np.linalg.norm(v_candidate)
                norm_h = np.linalg.norm(v_hist)
                
                if norm_c != 0 and norm_h != 0:
                    similarity = dot_product / (norm_c * norm_h)
                    if similarity >= threshold:
                        print(f"[Vector Dedup] Similarity too high: {similarity:.2f} >= {threshold} (comparing with: '{raw_text[:20]}...')")
                        return True
            except Exception as ve:
                print(f"[Warning] Cosine similarity error: {ve}")
        
        # Jaccard fallback
        similarity = get_caption_similarity(candidate_text, raw_text)
        if similarity >= jaccard_threshold:
            print(f"[Jaccard Dedup] Similarity too high: {similarity:.2f} >= {jaccard_threshold} (comparing with: '{raw_text[:20]}...')")
            return True
                
    return False

def save_post_to_db(caption, product_detail, slot, client):
    global API_ENABLED
    vec = None
    if API_ENABLED and client:
        vec = get_embedding(caption, client)
    blob = serialize_vector(vec) if vec else None
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        path_norm = __file__.replace("\\", "/").lower()
        platform = "x" if "x-bot" in path_norm else "facebook"
        cursor.execute(
            "INSERT INTO posts (platform, raw_text, topic, slot, embedding, timestamp) VALUES (?, ?, ?, ?, ?, ?)",
            (platform, caption, product_detail, slot, blob, int(time.time()))
        )
        conn.commit()
        conn.close()
        print("[DB] Successfully saved post memory.")
    except Exception as e:
        print(f"[DB] Save post failed: {e}")


def clean_overlay_text(text):
    if not text:
        return ""
    # Remove brackets
    text = re.sub(r'[【】\[\]\(\)\{\}「」『』〈〉《》〔〕〖〗㘀〙〚〛«»‹›]', '', text)
    # Remove emoji prefix/suffix and common symbols
    text = re.sub(r'^[•\-\*\d\.\s\u2013\|\s#@❤️🙏🚚🔥✨⚡🌟⚡️🎁🎗️🎀🎯📢📌📍🛒🛍️]+', '', text)
    text = re.sub(r'[•\-\*\s\u2013\|\s#@❤️🙏🚚🔥✨⚡🌟⚡️🎁🎗️🎀🎯📢📌📍🛒🛍️]+$', '', text)
    # Remove product prefix terms
    text = re.sub(r'^(?:สินค้าพร้อมส่ง|พร้อมส่ง|ของแท้|แท้|ลดล้างสต๊อก|ส่งด่วน|ส่งฟรี|🔥|ลดราคา)\s*', '', text)
    # Keep only alphanumeric, space, dot, comma, dash, and Thai script
    text = re.sub(r'[^\w\s\.\,\-\u0e00-\u0e7f]', '', text)
    # Strip .00 from prices/numbers in the text (e.g. 50.00 -> 50)
    text = re.sub(r'(\d+)\.00', r'\1', text)
    # Collapse double spaces
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def format_thai_price(price_val):
    if not price_val:
        return ""
    try:
        val = float(str(price_val).replace(",", "").strip())
        if val.is_integer():
            return str(int(val))
        else:
            return f"{val:.2f}".rstrip('0').rstrip('.')
    except Exception:
        s = str(price_val).strip()
        if s.endswith(".00"):
            s = s[:-3]
        return s


GEMINI_API_KEY    = os.environ.get("GEMINI_API_KEY",    "")
PAGE_ACCESS_TOKEN = os.environ.get("SOMTAM_PAGE_ACCESS_TOKEN", "")
PAGE_ID           = "554501167740603"
TEXT_MODELS       = ["gemini-flash-latest", "gemini-flash-latest"]
OUTPUT_DIR        = "output"
EXCEL_PATH        = os.path.join(os.path.dirname(__file__), "review_products.xlsx")
AFFILIATE_DIR     = os.path.join(os.path.dirname(__file__), "affiliate_data")
ACCENT_COLOR      = (255, 107, 53) # ส้ม #FF6B35 สำหรับพริก 10 เม็ด

if not GEMINI_API_KEY:
    try:
        from config import GEMINI_API_KEY, PAGE_ACCESS_TOKEN
    except ImportError:
        pass

if not GEMINI_API_KEY:
    GEMINI_API_KEY = "DUMMY_KEY"

API_ENABLED = True
client = None
if not GEMINI_API_KEY or GEMINI_API_KEY in ("DUMMY_KEY", "DUMMY"):
    print("[Warning] GEMINI_API_KEY is not set or is a dummy key. Disabling API calls.")
    API_ENABLED = False
else:
    try:
        client = genai.Client(api_key=GEMINI_API_KEY, http_options={'timeout': 60000.0})
    except Exception as e:
        print(f"[Warning] Failed to initialize genai.Client: {e}. Disabling API calls.")
        API_ENABLED = False

# --- Thai Helpers and Fallbacks ---
_LEADING_VOWELS  = set('เแโใไ')
_COMBINING_CHARS = set('่้๊๋์ิีึืุูัํ็')

THAI_WORDS = [
    "รายละเอียด", "โปรโมชั่น", "เครื่องมือ", "คอมพิวเตอร์", "แอปพลิเคชัน", "เก็บเงินปลายทาง",
    "โทรศัพท์", "แบตเตอรี่", "บัตรเครดิต", "พร้อมส่ง", "จัดส่ง", "ต่างประเทศ",
    "พรีออเดอร์", "ประหยัด", "ปลอดภัย", "คุ้มค่า", "สะดวกสบาย", "ธรรมชาติ",
    "คุณภาพ", "ภาพถ่าย", "พลาสติก", "ของแท้", "รับประกัน", "ลิขสิทธิ์",
    "แนะนำ", "สินค้า", "รีวิว", "สุดยอด", "ดีที่สุด", "สะดวก", "สบาย", "ง่ายดาย",
    "รวดเร็ว", "โปรโมชั่", "ส่วนลด", "คูปอง", "จัดส่ง", "ประกัน",
    "ชาร์จ", "หน้าจอ", "ลำโพง", "หูฟัง", "กล้อง", "เลนส์", "มือถือ", "ปุ่มกด",
    "สำหรับ", "เกี่ยวกับ", "อย่างไร", "เมื่อไหร่", "ที่ไหน", "เท่าไหร่",
    "ทุกคน", "ทุกวัน", "ทุกคืน", "สุดท้าย", "แรกเริ่ม", "จริงจัง",
    "สวัสดี", "ขอบคุณ", "ขอโทษ", "ยินดี", "หัวเราะ", "ร้องไห้",
    "ทำงาน", "พักผ่อน", "ออกกำลัง", "ท่องเที่ยว", "เดินทาง",
    "เก้าอี้", "โต๊ะทำงาน", "เบาะรอง", "พิงหลัง", "สายรัด", "การ์ตูน",
    "กระเป๋า", "รองเท้า", "เสื้อผ้า", "กางเกง", "นาฬิกา", "แว่นตา", "เครื่อง", "ระบบ",
    "ความสุข", "ร่างกาย", "สุขภาพ", "ออกกำลัง", "อาหาร", "ผลไม้", "น้ำดื่ม", "กาแฟ",
    "ราคา", "พิเศษ", "ทั่วไป", "ส่งฟรี", "ลดราคา", "ของแถม", "ปลายทาง",
    "ชั่วโมง", "นาที", "วินาที", "สัปดาห์", "ปีใหม่", "วันนี้", "พรุ่งนี้", "เมื่อวาน",
    "ใครก็ตาม", "สิ่งใด", "ทั้งหมด", "บางส่วน", "ประเภท", "รูปแบบ",
    "ติดตาม", "กดไลก์", "แชร์โพส", "คอมเมนต์", "คลิกลิงก์", "พิกัด", "ชี้เป้า",
    "ค่ะ", "ครับ", "ผม", "เรา", "คุณ", "ท่าน",
    "พี่", "น้อง", "พ่อ", "แม่", "เพื่อน", "บ้าน", "เมือง", "เวลา", "ดีใจ", "เสียใจ", 
    "รัก", "ชอบ", "เกลียด", "กลัว", "โกรธ", "ทำ", "กิน", "นอน", "เดิน", "วิ่ง", "นั่ง", 
    "ยืน", "พูด", "ฟัง", "ดู", "เห็น", "คิด", "รู้", "จำ", "ลืม", "เรียน", "เล่น", "ซื้อ", 
    "ขาย", "ราคา", "ถูก", "แพง", "ลด", "แถม", "ส่ง", "ด่วน", "ฟรี", "รับ", "ศูนย์",
    "แท้", "ใหม่", "เก่า", "แรก", "นี้", "นั้น", "โน้น", "นี่", "นั่น", "โน่น", "อะไร", 
    "ใคร", "กี่", "บ้าง", "ทุก", "บาง", "จริง", "จัง", "แท้", "เทียม", "ปลอม", "สาย", 
    "เคส", "ฟิล์ม", "ภาพ", "รูป", "เสียง", "เพลง", "หนัง", "เกม", "แอป", "เว็บ", "เน็ต", 
    "โค้ด", "โอน", "หวย", "ออก", "เงิน", "เก็บ", "แสน", "แรก", "งาน", "การ", "ช่วย", 
    "บอก", "ให้", "คน", "ทอง", "ร้อย", "พัน", "หมื่น", "ล้าน", "มาก", "น้อย", "ดี", 
    "เลว", "ชั่ว", "สูง", "ต่ำ", "ดำ", "ขาว", "แดง", "เขียว", "เหลือง", "ฟ้า", "ส้ม", 
    "ชมพู", "ม่วง", "เทา", "สวย", "หล่อ", "และ", "หรือ", "แต่", "ที่", "ซึ่ง", "อัน", 
    "ของ", "เพื่อ", "ใน", "จาก", "โดย", "ตาม", "กับ", "มี", "เป็น", "จะ", "ต้อง", 
    "อยาก", "นุ่ม", "แข็ง", "ใหญ่", "เล็ก", "ยาว", "สั้น", "กว้าง", "แคบ", "หนา", 
    "บาง", "ร้อน", "เย็น", "อุ่น", "หนาว", "ง่าย", "ยาก", "เร็ว", "ช้า", "ได้", 
    "เลย", "ด้วย", "จาก", "ถึง", "จน", "กว่า", "ก็", "ยัง", "อีก", "แล้ว", "นะ", 
    "สิ", "ละ", "หน่อย", "นิด", "ชิ้น", "กล่อง", "อัน", "ตัว", "ใบ", "คู่", "ชุด", 
    "แผ่น", "ม้วน"
]

def contains_thai(text):
    if not text:
        return False
    return bool(re.search(r'[\u0e00-\u0e7f]', text))

def local_segment_thai(text):
    if not text:
        return ""
    word_set = set(THAI_WORDS)
    max_len = max(len(w) for w in THAI_WORDS)
    
    result = []
    i = 0
    n = len(text)
    
    while i < n:
        if not contains_thai(text[i]):
            result.append(text[i])
            i += 1
            continue
            
        matched = False
        for l in range(min(max_len, n - i), 0, -1):
            substr = text[i:i+l]
            if substr in word_set:
                result.append(substr)
                i += l
                matched = True
                break
        
        if not matched:
            start = i
            while i < n and contains_thai(text[i]):
                word_matched_here = False
                if i > start:
                    for l in range(min(max_len, n - i), 0, -1):
                        if text[i:i+l] in word_set:
                            word_matched_here = True
                            break
                if word_matched_here:
                    break
                i += 1
            result.append(text[start:i])
            
    output = []
    for idx, part in enumerate(result):
        if idx > 0:
            prev_char = result[idx-1][-1]
            curr_char = part[0]
            if (contains_thai(prev_char) and contains_thai(curr_char) and 
                prev_char != '\u200b' and curr_char != '\u200b' and
                curr_char not in _COMBINING_CHARS and
                prev_char not in _LEADING_VOWELS):
                output.append('\u200b')
        output.append(part)
        
    return "".join(output)

def segment_thai_text(text, client=client):
    global API_ENABLED
    if not text or not contains_thai(text):
        return text
    if not API_ENABLED or client is None:
        return local_segment_thai(text)
    prompt = (
        "You are an expert Thai word segmentation tool. "
        "Your task is to insert a zero-width space character (\\u200b) at every natural word boundary in the provided Thai text. "
        "Strict rules:\n"
        "1. Do NOT modify, delete, or add any words, characters, punctuation, spaces, or newlines of the original text. "
        "Keep the exact same characters and layout.\n"
        "2. Do NOT add any introductory or concluding remarks. Output ONLY the segmented text.\n"
        "3. Ensure words like 'หวยออก', 'เงินเก็บ', 'แสนแรก', 'ทำงาน' are segmented at their natural boundaries (e.g., 'หวย\\u200bออก' or left as 'หวยออก', but never break syllables awkwardly).\n\n"
        f"Text to segment:\n{text}"
    )
    for model_idx, model in enumerate(TEXT_MODELS):
        if model_idx > 0:
            time.sleep(2)
        try:
            resp = client.models.generate_content(model=model, contents=prompt)
            segmented = resp.text.strip().replace('\\u200b', '\u200b')
            clean_orig = text.replace('\u200b', '').replace('\\u200b', '')
            clean_seg = segmented.replace('\u200b', '').replace('\\u200b', '')
            if len(clean_orig) == len(clean_seg):
                return segmented
        except Exception as e:
            print(f"[{model}] segment_thai_text failed: {e}")
            
    print("[Warning] segment_thai_text failed on all models. Disabling API calls for this run.")
    API_ENABLED = False
    return local_segment_thai(text)

def load_next_product():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=False):
        no    = row[0].value
        detail= row[1].value
        shopee= row[2].value
        lazada= row[3].value
        imgurl= row[4].value
        promo = row[5].value
        posted= row[6].value

        # ข้าม sample row และ posted แล้ว
        if not detail or "วางรายละเอียด" in str(detail):
            continue
        if str(posted).strip().lower() == "done" or str(posted).strip().startswith("done"):
            continue
        if not shopee or "xxx" in str(shopee):
            continue
        if not imgurl or "xxx" in str(imgurl) or not str(imgurl).strip().startswith("http"):
            continue

        return {
            "row": row[0].row,
            "no": no,
            "detail": str(detail).strip(),
            "shopee": str(shopee).strip(),
            "lazada": str(lazada).strip() if lazada else "",
            "image_url": str(imgurl).strip() if imgurl else "",
            "promo": str(promo).strip() if promo else "",
        }, wb, ws
    return None, wb, ws


# --- Upgraded Variety & Quality Upgrades (Memory, Roles, Specs, Similarity) ---

ROLES = [
    {"id": "R1", "name": "พ่อบ้าน", "desc": "พ่อบ้านวัยทำงาน ชอบจัดบ้าน ทำงานบ้าน และดูแลความเรียบร้อยในบ้าน"},
    {"id": "R2", "name": "สายกิน", "desc": "คนชอบกิน ชอบลองของอร่อย รีวิวของกินเป็นหลัก หรือเน้นความคุ้มค่าด้านอาหาร"},
    {"id": "R3", "name": "คนลดน้ำหนัก", "desc": "คนกำลังควบคุมอาหาร ดูแลรูปร่าง สุขภาพ และแคลอรี่"},
    {"id": "R4", "name": "มนุษย์เงินเดือน", "desc": "พนักงานออฟฟิศ/คนวัยทำงานที่ยุ่งตลอดวัน มักหาความสะดวกสบาย อุปกรณ์ออฟฟิศ หรือขนมยามบ่าย"},
    {"id": "R5", "name": "คนเลี้ยงสัตว์", "desc": "คนรักสัตว์ เลี้ยงหมา เลี้ยงแมว รักพวกมันเหมือนลูก"},
    {"id": "R6", "name": "นักออกกำลังกาย", "desc": "สายฟิตเนส คนรักสุขภาพ ออกกำลังกายเป็นประจำ สนใจโปรตีนและอุปกรณ์กีฬา"},
    {"id": "R7", "name": "แม่บ้าน", "desc": "แม่บ้านครอบครัวที่ใส่ใจเรื่องความประหยัด ความสะอาด ความสะดวกสบายของคนในบ้าน"},
    {"id": "R8", "name": "คนชอบของถูก", "desc": "นักช้อปสายประหยัด ชอบหารเฉลี่ย ราคาต่อชิ้น ความคุ้มค่าและโปรลดราคา"}
]

def choose_role(state):
    last_roles = state.get("last_roles", [])
    forbidden = []
    if len(last_roles) >= 2 and last_roles[-1] == last_roles[-2]:
        forbidden.append(last_roles[-1])
    available = [r for r in ROLES if r["id"] not in forbidden]
    if not available:
        available = ROLES
    return random.choice(available)

def get_caption_similarity(cap1, cap2):
    def normalize_text(text):
        text = text.lower()
        text = re.sub(r'[^\w\u0e00-\u0e7f]', '', text)
        return text

    n1 = normalize_text(cap1)
    n2 = normalize_text(cap2)
    if not n1 or not n2:
        return 0.0
    def get_bigrams(text):
        return set(text[i:i+2] for i in range(len(text)-1))
    b1 = get_bigrams(n1)
    b2 = get_bigrams(n2)
    if not b1 or not b2:
        return 0.0
    intersection = b1.intersection(b2)
    union = b1.union(b2)
    return len(intersection) / len(union)

def clean_product_title_locally(title):
    # Strip specs and sizes
    title = re.sub(r'\b\d+\s*(?:[lL]|[mM][lL]|[gG]|[kK][gG]|[bB][tT][uU]|[pP][cC][sS]|[pP][aA][cC][kK]|\s*กล่อง|\s*ซอง|\s*แพ็ค|\s*ตัว|\s*คู่)\b', '', title)
    title = re.sub(r'\d+\s*/\s*\d+', '', title)
    title = re.sub(r'/\s*/', '/', title)
    title = re.sub(r'\b\d+in\d+\b', '', title)
    title = re.sub(r'\b(?:PM\d+\.?\d*|\d+|[a-zA-Z]+\d+|\d+[a-zA-Z]+)\b', '', title)
    title = re.sub(r'\b[a-zA-Z\d]{2,10}\d[a-zA-Z\d]*\b', '', title)
    title = re.sub(r'[【】\[\]\(\)\{\}「」『』〈〉《》〔〕〖〗〘〙〚〛«»‹›]', '', title)
    title = re.sub(r'^(?:สินค้าขายดี|สินค้าพร้อมส่ง|พร้อมส่ง|ของแท้|แท้|ลดล้างสต๊อก|ส่งด่วน|ส่งฟรี|ลดราคา|จัดส่งฟรี)\s*', '', title)
    title = re.sub(r'\s+', ' ', title).strip()
    words = title.split()
    if len(words) > 4:
        title = " ".join(words[:4])
    return title.strip()

def get_role_context(role_name, prod_type):
    phrases = {
        "พ่อบ้าน": "ในฐานะพ่อบ้านที่ชอบจัดบ้านดูแลงานบ้าน บอกเลยว่าถูกใจมาก",
        "สายกิน": "สายของอร่อยอย่างเราลองแล้วดื่มด่ำมาก ฟินสุดๆ",
        "คนลดน้ำหนัก": "ช่วงควบคุมอาหาร ดูแลหุ่นอยู่ ได้ตัวนี้ช่วยชีวิตไว้เลย แคลน้อยโปรตีนสูง",
        "มนุษย์เงินเดือน": "แช่ตู้เย็นที่ออฟฟิศไว้ เวลาหิวดึกหิวยามบ่ายช่วยชีวิตพนักงานออฟฟิศได้เยอะ",
        "คนเลี้ยงสัตว์": "คนรักสัตว์อย่างเราหาของดีๆ มาให้เด็กๆ ตลอด ตัวนี้ตอบโจทย์มาก",
        "นักออกกำลังกาย": "หลังเวทเสร็จเหนื่อยๆ ขี้เกียจเคี้ยวอกไก่ ได้ตัวนี้ดื่มทดแทนสะดวกมาก",
        "แม่บ้าน": "ในฐานะแม่บ้านที่ชอบหาของเข้าครัว ดูแลคนในบ้าน แนะนำเลยค่ะ",
        "คนชอบของถูก": "สายช้อปของคุ้มหารเฉลี่ยราคาต่อชิ้นแล้วประหยัดมาก คุ้มค่าน้ำตาไหล"
    }
    prod_type_lower = prod_type.lower()
    if "นม" in prod_type_lower or "โปรตีน" in prod_type_lower or "เวย์" in prod_type_lower:
        phrases["นักออกกำลังกาย"] = "หลังออกกำลังกายเสร็จเหนื่อยๆ บางวันเคี้ยวอกไก่ไม่ไหวจริงๆ ได้ตัวนี้ดื่มแทนง่ายมาก"
        phrases["มนุษย์เงินเดือน"] = "พกไปกินตอนบ่ายที่ออฟฟิศอิ่มท้องกำลังดี ไม่ต้องเดินไปกดน้ำหวาน"
        phrases["คนลดน้ำหนัก"] = "ช่วยลดยากินน้ำหวานจุกจิกได้เยอะ แคลคุมง่ายอิ่มท้องนาน"
    elif any(x in prod_type_lower for x in ["สัตว์", "แมว", "หมา", "สุนัข", "cat", "dog", "pet"]):
        phrases["คนเลี้ยงสัตว์"] = "เด็กๆ ที่บ้านชอบกันใหญ่ แย่งกันกินตึม สรรหามานานในที่สุดก็เจอ"
    return phrases.get(role_name, "ส่วนตัวลองใช้แล้วประทับใจมาก")

PERSONAS = [
    {"id": "A", "name": "เพื่อนบอกต่อ", "desc": "เพื่อนบอกต่อ แนะนำของดีให้เพื่อนด้วยความเป็นกันเอง"},
    {"id": "B", "name": "คนซื้อมาใช้แล้ว", "desc": "คนที่ชอบซื้อของออนไลน์และมารีวิวสั้นๆ หลังใช้งานจริงมาระยะหนึ่ง บอกเล่าตามจริง"},
    {"id": "C", "name": "คนเจอโปรมาแชร์", "desc": "คนที่บังเอิญเจอราคาโปรโมชั่นหรือส่วนลดพิเศษแล้วอยากเอามาแชร์ต่อด้วยความตื่นเต้นเบาๆ"},
    {"id": "D", "name": "คนเคยมีปัญหาแล้วสินค้าช่วยแก้", "desc": "คนที่มีปัญหาชีวิตประจำวันแล้วสินค้าตัวนี้เข้ามาช่วยแก้ปัญหาได้ตรงจุด"},
    {"id": "E", "name": "คนเล่าเรื่องชีวิตประจำวัน", "desc": "คนเล่าเรื่องราวชีวิตประจำวัน/อุปสรรคชีวิตทั่วไปก่อน แล้วโยงเข้าหาตัวสินค้าที่เข้ามาช่วยชีวิต"},
    {"id": "F", "name": "คนรีวิวสั้นๆแบบขำๆ", "desc": "คนรีวิวสั้นๆ แบบขำๆ ตลกๆ สอดแทรกความตลกหรือมุกแซวตัวเอง"},
    {"id": "G", "name": "คนสงสัยก่อนซื้อแล้วมาลอง", "desc": "คนที่เคยสงสัย/ไม่แน่ใจในคุณภาพของสินค้าตัวนี้มาก่อน แล้วตัดสินใจซื้อมาลองและมาแชร์ความจริง"},
    {"id": "H", "name": "คนแนะนำของให้ครอบครัว", "desc": "คนที่ห่วงใยครอบครัวหรือคนใกล้ตัว แนะนำของดีให้คนในบ้าน/ครอบครัวได้ใช้"}
]

HOOKS = [
    "เมื่อวานเพิ่งเจอ...",
    "ไม่คิดว่าจะ...",
    "อันนี้เกินคาดจริงๆ",
    "มีใครเป็นเหมือนผมไหม",
    "ตอนแรกไม่ได้จะซื้อ",
    "กำลังหาของแบบนี้อยู่พอดี",
    "เห็นคนพูดถึงเยอะเลยลอง",
    "โดนป้ายยามาอีกที",
    "ลองแล้วเข้าใจเลยว่าทำไมขายดี",
    "แชร์เผื่อมีคนกำลังหาอยู่"
]

STYLES = [
    {"name": "มุกตลก", "desc": "มีมุกตลกหรือการแซวตัวเอง ขำขัน"},
    {"name": "ประสบการณ์ส่วนตัว (PAS)", "desc": "เขียนด้วยโครงสร้าง PAS: ชี้ปัญหาความลำบากในชีวิตประจำวัน (Problem) -> ขยี้ขยายความรู้สึกอึดอัดทรมาน (Agitate) -> แนะนำสินค้าเป็นทางออกที่เข้ามาช่วยชีวิต (Solution)"},
    {"name": "คำถาม", "desc": "เปิดประเด็นด้วยคำถาม ชวนคุยหรือชวนคิด"},
    {"name": "การเปรียบเทียบ (BAB)", "desc": "เขียนด้วยโครงสร้าง BAB: เล่าชีวิตช่วงที่แย่เมื่อก่อน (Before) -> เล่าชีวิตที่ดีขึ้นเมื่อใช้สินค้าแล้ว (After) -> เชื่อมโยงว่าสินค้าตัวนี้เปลี่ยนชีวิตอย่างไร (Bridge)"},
    {"name": "การบอกต่อเฉยๆ", "desc": "เป็นการบอกต่อเฉยๆ แบบป้ายยาเพื่อนธรรมดา เป็นกันเอง"}
]

def load_state(wb):
    state = {
        "last_personas": [],
        "last_hooks": [],
        "last_styles": [],
        "last_roles": [],
        "recent_captions": []
    }
    if "posted_state" in wb.sheetnames:
        ws = wb["posted_state"]
        val = ws.cell(row=1, column=1).value
        if val:
            try:
                import json
                loaded = json.loads(val)
                state.update(loaded)
            except Exception as e:
                print(f"[State] Failed to parse state JSON: {e}")
    return state

def save_state_to_wb(wb, state):
    if "posted_state" not in wb.sheetnames:
        ws = wb.create_sheet("posted_state")
    else:
        ws = wb["posted_state"]
    import json
    state["last_personas"] = state.get("last_personas", [])[-50:]
    state["last_hooks"] = state.get("last_hooks", [])[-50:]
    state["last_styles"] = state.get("last_styles", [])[-50:]
    state["last_roles"] = state.get("last_roles", [])[-50:]
    state["recent_captions"] = state.get("recent_captions", [])[-50:]
    ws.cell(row=1, column=1, value=json.dumps(state, ensure_ascii=False))

def choose_persona(state):
    last_personas = state.get("last_personas", [])
    forbidden = []
    if len(last_personas) >= 2 and last_personas[-1] == last_personas[-2]:
        forbidden.append(last_personas[-1])
    available = [p for p in PERSONAS if p["id"] not in forbidden]
    if not available:
        available = PERSONAS
    return random.choice(available)

def choose_hook(state):
    last_hooks = state.get("last_hooks", [])
    window_size = min(20, len(HOOKS) - 1)
    forbidden = last_hooks[-window_size:] if last_hooks else []
    available = [h for h in HOOKS if h not in forbidden]
    if not available:
        available = HOOKS
    return random.choice(available)

def choose_style(state):
    last_styles = state.get("last_styles", [])
    forbidden = []
    if len(last_styles) >= 2 and last_styles[-1] == last_styles[-2]:
        forbidden.append(last_styles[-1])
    available = [s for s in STYLES if s["name"] not in forbidden]
    if not available:
        available = STYLES
    return random.choice(available)

def mark_posted(wb, ws, row_num, state=None):
    bkk = timezone(timedelta(hours=7))
    ts = datetime.now(bkk).strftime("%Y-%m-%d %H:%M")
    ws.cell(row=row_num, column=7, value=f"done {ts}")
    if state is not None:
        save_state_to_wb(wb, state)
    wb.save(EXCEL_PATH)
    print(f"Marked row {row_num} as done")

def get_allowed_xlsx_files():
    path = os.path.abspath(__file__).replace("\\", "/")
    if "chowchow" in path:
        return ["สัตว์เลี้ยง.xlsx"]
    elif "somtam" in path:
        return ["อาหารและเครื่องดื่ม.xlsx"]
    elif "rocket" in path:
        return ["เครื่องใช้ไฟฟ้าภายในบ้าน.xlsx"]
    elif "x-bot" in path:
        return ["สินค้าขายดี.xlsx"]
    else:  # kram-facebook-page
        return ["เครื่องใช้ในบ้าน.xlsx", "ค่าคอมพิเศษ.xlsx", "เสื้อผ้าแฟชั่นผู้หญิง.xlsx", "สินค้าสำหรับเม้นใต้คลิป.xlsx"]

def get_posted_urls(ws):
    urls = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 4:
            continue
        shopee = row[2]
        lazada = row[3]
        if shopee and str(shopee).strip().startswith("http"):
            urls.add(str(shopee).strip())
        if lazada and str(lazada).strip().startswith("http"):
            urls.add(str(lazada).strip())
    return urls

def append_posted_fallback(wb, ws, product, state=None):
    bkk = timezone(timedelta(hours=7))
    ts = datetime.now(bkk).strftime("%Y-%m-%d %H:%M")
    row_num = ws.max_row + 1
    
    ws.cell(row=row_num, column=1, value=product["no"])
    ws.cell(row=row_num, column=2, value=product["detail"])
    ws.cell(row=row_num, column=3, value=product["shopee"])
    ws.cell(row=row_num, column=4, value=product["lazada"])
    ws.cell(row=row_num, column=5, value=product["image_url"])
    ws.cell(row=row_num, column=6, value=product["promo"])
    ws.cell(row=row_num, column=7, value=f"done {ts}")
    
    if state is not None:
        save_state_to_wb(wb, state)
    wb.save(EXCEL_PATH)
    print(f"Appended fallback product to review_products.xlsx at row {row_num}")

def load_affiliate_product(posted_urls):
    """Fallback: สุ่มสินค้าจาก AFFILIATE_DIR เมื่อ review_products.xlsx หมด"""
    import glob
    allowed_names = get_allowed_xlsx_files()
    xlsx_files = []
    for name in allowed_names:
        p = os.path.join(AFFILIATE_DIR, name)
        if os.path.exists(p):
            xlsx_files.append(p)
            
    # Fallback to any xlsx files if no mapped files exist
    if not xlsx_files:
        xlsx_files = glob.glob(os.path.join(AFFILIATE_DIR, "*.xlsx"))
        
    if not xlsx_files:
        print(f"[affiliate] No xlsx files found in {AFFILIATE_DIR}")
        return None
        
    random.shuffle(xlsx_files)
    for xlsx_path in xlsx_files:
        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb.active
            candidates = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 10:
                    continue
                name   = row[1]
                imgurl = row[2]
                price  = row[3]
                shopee = row[9]
                lazada = row[10] if len(row) > 10 else None
                if not shopee or not name:
                    continue
                
                shopee_val = str(shopee).strip()
                lazada_val = str(lazada).strip() if lazada else ""
                
                # Check against posted URLs
                if shopee_val in posted_urls or (lazada_val and lazada_val in posted_urls):
                    continue
                    
                candidates.append({
                    "no": row[0],
                    "detail": f"{name} ราคา {price} บาท",
                    "shopee": shopee_val,
                    "lazada": lazada_val,
                    "image_url": str(imgurl).strip() if imgurl else "",
                    "promo": "",
                    "row": None,
                })
            wb.close()
            if candidates:
                product = random.choice(candidates)
                print(f"[affiliate] Loaded from {os.path.basename(xlsx_path)}: {product['detail'][:60]}")
                return product
        except Exception as e:
            print(f"[affiliate] Failed to read {xlsx_path}: {e}")
    print("[affiliate] No valid product found in any xlsx file")
    return None


def clean_promo(raw):
    """เอาเฉพาะบรรทัดที่มี ฿ หรือ ลด หรือ % หรือ ส่งฟรี"""
    if not raw:
        return ""
    lines = raw.strip().splitlines()
    kept = [l.strip() for l in lines if re.search(r'฿|ลด|%|ส่งฟรี|flash|sale', l, re.IGNORECASE)]
    return " | ".join(kept[:3]) if kept else ""

def extract_highlights(detail, promo):
    """ให้ AI สกัดจุดเด่นจาก raw detail"""
    global API_ENABLED
    highlights = None
    if API_ENABLED:
        prompt = (
            f"จากรายละเอียดสินค้านี้:\n{detail}\n\n"
            f"สกัดจุดเด่นสินค้าเป็นประโยคข้อความสั้นแนวธรรมชาติ 1-2 ย่อหน้าสั้นๆ (ห้ามทำเป็นข้อๆ หรือมีสัญลักษณ์รายการ/bullet points เช่น •, ▪️, - หรือเลขข้อ) "
            f"เน้นประโยชน์ที่คนซื้อสนใจและใช้งานจริง ห้ามใส่ข้อมูลราคาหรือโปรโมชั่น "
            f"ตอบเฉพาะส่วนรายละเอียดเนื้อความเท่านั้น"
        )
        for model_idx, model in enumerate(TEXT_MODELS):
            if model_idx > 0:
                time.sleep(2)
            try:
                resp = client.models.generate_content(model=model, contents=prompt)
                highlights = resp.text.strip()
                if highlights:
                    break
            except Exception as e:
                print(f"[{model}] highlights failed: {str(e)[:80]}")
                
        if not highlights:
            print("[Warning] Highlights generation failed on all models. Disabling API calls for this run.")
            API_ENABLED = False
            
    if not highlights:
        print("[Warning] Falling back to local heuristic extraction.")
        lines = [l.strip() for l in detail.splitlines() if l.strip()]
        thai_lines = [l for l in lines if contains_thai(l)]
        if not thai_lines:
            thai_lines = lines
        
        points = []
        for line in thai_lines:
            cleaned = re.sub(r'^[•\-\*\d\.\s\u2013]+', '', line).strip()
            if cleaned and len(cleaned) > 5 and len(cleaned) < 100:
                points.append(cleaned)
            if len(points) >= 3:
                break
        if not points:
            points = [line[:80] for line in thai_lines[:2]]
        highlights = " ".join(points) if points else "รายละเอียดเพิ่มเติมศึกษาต่อได้ที่หน้าร้านเลยค่ะ"
        
    if promo:
        highlights += f"\n🔥 โปรตอนนี้: {promo}"
    return highlights

def download_image(url):
    """Download รูปแรกหรือสุ่มรูปจาก gallery"""
    urls = [u.strip() for u in url.strip().split("\n") if u.strip()]
    if not urls:
        raise RuntimeError("No image URLs provided")
    if len(urls) > 1:
        import random
        if random.random() < 0.40:
            chosen_url = urls[0]
        else:
            chosen_url = random.choice(urls[1:])
    else:
        chosen_url = urls[0]
    resp = requests.get(chosen_url, timeout=15, headers={"User-Agent": "Mozilla/5.0"})
    if resp.status_code == 200:
        bkk = timezone(timedelta(hours=7))
        ts = datetime.now(bkk).strftime("%Y%m%d_%H%M%S")
        ext = "webp" if "webp" in chosen_url else "jpg"
        path = os.path.join(OUTPUT_DIR, f"product_{ts}.{ext}")
        with open(path, "wb") as f:
            f.write(resp.content)
        print(f"Product image downloaded: {path}")
        return path
    raise RuntimeError(f"Image download failed: {resp.status_code}")

def generate_hook(detail, highlights):
    """สร้างพาดหัวการ์ดรีวิวสไตล์ CatDumb ภาษาพูดคนไทยธรรมชาติ 2 บรรทัด คั่นด้วย '|'"""
    global API_ENABLED
    active_client = globals().get('client')
    if API_ENABLED and active_client is not None:
        prompt = (
            "คุณคือ Copywriter มืออาชีพสไตล์ CatDumb (แคทดัมบ์) ภาษาพูดคนไทยธรรมดา ธรรมชาติ สนุก เป็นกันเอง ชี้เป้าของน่าซื้อ\n"
            "จงสร้างพาดหัวรีวิวสินค้าภาษาไทย 2 บรรทัด คั่นด้วย '|' (บรรทัด 1 | บรรทัด 2)\n\n"
            "กฎเหล็ก:\n"
            "1. บรรทัดที่ 1 (ฉายาเด็ด/ชื่อโปรย): ตั้งชื่อเล่นสั้นๆ สไตล์พูดปากต่อปาก (2-4 คำ) สะดุดตา ไม่เอาชื่อรุ่น/ยี่ห้อ/สเปกแข็งๆ เด็ดขาด เช่น 'ธัญพืช 9 เซียนสายคลีน', 'แว่นกันแดดทรงในตำนาน', 'แก้วเก็บเย็นข้ามวัน', 'เบาะรองนั่งสู้ชีวิต'\n"
            "2. บรรทัดที่ 2 (จุดเด่นภาษาพูด): สรุปจุดเด่นหรือประโยชน์ที่ได้รับเป็นภาษาพูดธรรมดา (3-6 คำ) ไม่ใช้ภาษาสเปกทางการ ไม่ยาวเกินไป เช่น 'ไม่ใส่เนยน้ำตาล อบใหม่รสธรรมชาติ', 'นั่งทำงาน 10 ชม. ไม่ปวดเอว', 'ฟังเพลง 8 ชม. เบสแน่นตึ้บ'\n"
            "3. ห้ามใส่เครื่องหมายคำพูด (\"\") ห้ามมี Emoji ห้ามมีคำว่า คุ้มค่า, คุ้มสุดๆ, ห้ามพลาด, ดีงาม, ต้องจัด, ของดี\n"
            "4. เว้นวรรคเฉพาะจังหวะภาษาไทยปกติ ห้ามเว้นวรรคแตกคำย่อย\n\n"
            f"ข้อมูลสินค้า:\n{detail}\n\n"
            f"จุดเด่นสินค้า:\n{highlights}\n\n"
            "ผลลัพธ์ (บรรทัด 1 | บรรทัด 2):"
        )
        for model_idx, model in enumerate(TEXT_MODELS):
            if model_idx > 0:
                time.sleep(2)
            try:
                resp = active_client.models.generate_content(model=model, contents=prompt)
                result = resp.text.strip()
                label_pattern = r'^(ข้อความในโพสต์\s*Facebook|Facebook\s*Caption|Facebook\s*caption|Caption|caption|ข้อความบนรูป|ข้อความในรูป|ข้อความ|คำบรรยาย|คำอธิบาย|บรรทัดที่\s*\d+|บรรทัด\s*\d+|ประโยคที่\s*\d+|ประโยค\s*\d+|Hook\s*text|Hook|Line\s*\d+|[L|l]ine\s*\d+|\d+)\s*[:\-\.\s]\s*'
                result = re.sub(label_pattern, '', result, flags=re.IGNORECASE).strip()
                result = result.strip('"\'“”‘’')
                if "|" in result:
                    parts = result.split("|", 1)
                    line1 = parts[0].strip()
                    line2 = parts[1].strip()
                    return line1, line2
                else:
                    return result[:15], ""
            except Exception as e:
                print(f"[{model}] hook generation failed: {e}")
        print("[Warning] Hook generation failed on all models. Disabling API calls for this run.")
        API_ENABLED = False
    # Local fallback for hook
    title = re.sub(r'^[•\-\*\d\.\s\u2013\(\[\{\)\|\}]+', '', detail).strip()
    first_line = title.split('\n')[0].split('|')[0].split(' - ')[0].split(' – ')[0].strip()
    # line1: ตัดชื่อสั้นก่อน spec keyword แล้วเอา 3 คำแรก
    first_line = re.sub(r'[\[\]\(\)]', '', first_line)
    _specs = ['ราคา', 'แพ็ค', 'ขนาด', 'สำหรับ', 'จำนวน', 'กรัม', 'ลิตร', ' ml', ' kg', ' g ', ',']
    _stop = len(first_line)
    for _kw in _specs:
        _idx = first_line.find(_kw)
        if _idx > 2:
            _stop = min(_stop, _idx)
    short_title = first_line[:_stop].strip()
    words = short_title.split()
    line1 = "สินค้าแนะนำ"
    for n in (3, 2, 1):
        candidate = " ".join(words[:n]) if words else "สินค้าแนะนำ"
        if len(candidate) <= 20 or n == 1:
            line1 = candidate
            break
    # line2: ราคา ถ้ามี / ไม่มีดึง feature แรก
    price_m = re.search(r'ราคา\s*([\d,]+(?:\.\d+)?)\s*บาท', detail)
    if price_m:
        line2 = f"ราคา {format_thai_price(price_m.group(1))} บาท"
    else:
        all_lines = [l.strip() for l in detail.splitlines() if l.strip() and len(l.strip()) > 8]
        clean_lines = [re.sub(r'^[^฀-๿\w]+', '', l).strip() for l in all_lines]
        clean_lines = [l for l in clean_lines if l and not l.startswith('http')]
        line2 = " ".join(clean_lines[1].split()[:5]) if len(clean_lines) > 1 else ""
    return line1, line2

def local_parse_detail_to_json(detail, promo=None):
    price = ""
    # Find price in detail or promo
    price_m = re.search(r'(?:ราคา|฿)\s*([\d,]+(?:\.\d+)?)\s*(?:บาท)?', detail + " " + (promo or ""))
    if price_m:
        price = format_thai_price(price_m.group(1))
    
    lines = [l.strip() for l in detail.splitlines() if l.strip()]
    first_line = lines[0] if lines else ""
    first_line = re.sub(r'^[•\-\*\d\.\s\u2013\(\[\{\)\|\}#@❤️🙏🚚🔥]+', '', first_line).strip()
    clean_title = clean_product_title_locally(first_line)
    # Clean price from clean_title
    clean_title = re.sub(r'(?:ราคา|฿)\s*[\d,]+(?:\.\d+)?\s*(?:บาท)?', '', clean_title).strip()
    
    prod_type = "สินค้า"
    title_lower = clean_title.lower()
    
    if "ขนม" in title_lower:
        prod_type = "ของกิน"
    elif any(x in title_lower for x in ["กรงแมว", "อาหารเปียก", "อาหารแมว", "อาหารสุนัข", "ทรายแมว", "สัตว์เลี้ยง", "แมว", "หมา", "สุนัข", "cat", "dog", "pet"]):
        prod_type = "ของใช้สัตว์เลี้ยง"
    elif any(x in title_lower for x in ["นม", "โปรตีน", "เวย์", "ไฮโปรตีน"]) and "อัตโนมัติ" not in title_lower and "มอเตอร์" not in title_lower:
        prod_type = "นมโปรตีน"
    elif any(x in title_lower for x in ["ทุเรียน", "มะพร้าว", "เก๊กฮวย", "ชาเขียว", "มัทฉะ", "น้ำพริก", "คุกกี้", "เบเกอรี่", "กาแฟ", "ช็อกโกแลต", "อาหาร", "กิน", "ส้มตำ", "ปลา"]):
        prod_type = "ของกิน"
    elif any(x in title_lower for x in ["แว่น", "ray-ban", "rayban", "กันแดด"]):
        prod_type = "แว่นตากันแดด"
    elif any(x in title_lower for x in ["ทีที", "tv", "ทีวี", "โทรทัศน์", "smart tv"]):
        prod_type = "ทีวี"
    elif any(x in title_lower for x in ["เก้าอี้", "chair", "ergonomic"]):
        prod_type = "เก้าอี้สุขภาพ"
    elif any(x in title_lower for x in ["แก้ว", "กระติก", "tumbler"]):
        prod_type = "แก้วเก็บความเย็น"
    elif any(x in title_lower for x in ["หมอน", "ที่นอน", "เตียง", "pillow"]):
        prod_type = "เครื่องนอน"
    elif any(x in title_lower for x in ["ตัดหญ้า", "เครื่องมือ"]):
        prod_type = "เครื่องมือช่าง/ทำสวน"
        
    highlights = []
    for line in lines[1:]:
        cleaned = re.sub(r'^[•\-\*\d\.\s\u2013\(\[\{\)\|\}]+', '', line).strip()
        cleaned = re.sub(r'(?:ราคา|฿)\s*[\d,]+(?:\.\d+)?\s*(?:บาท)?', '', cleaned).strip()
        if cleaned and 5 < len(cleaned) < 80 and not cleaned.startswith("http"):
            highlights.append(cleaned)
        if len(highlights) >= 2:
            break
            
    if not highlights:
        highlights = [clean_title[:45]]
        
    res = {
        "ประเภท": prod_type,
        "จุดเด่น": ", ".join(highlights),
    }
    if price:
        res["ราคา"] = price
        
    flavors = []
    for f in ["ช็อกโกแลต", "สตอเบอรี่", "สตรอว์เบอร์รี่", "มะพร้าว", "ชาเขียว", "ทุเรียนหมอนทอง", "รสหวาน", "จืด", "รสเผ็ด", "ต้มยำ"]:
        if f in detail:
            flavors.append(f)
    if flavors:
        res["รสชาติ"] = ", ".join(flavors)
        
    colors = []
    for c in ["สีดำ", "สีขาว", "สีครีม", "สีเขียว", "สีแดง", "สีน้ำเงิน", "สีเทา", "black", "white", "grey"]:
        if c in detail.lower():
            colors.append(c)
    if colors:
        res["สี"] = ", ".join(colors)
        
    return res

def parse_detail_to_json(detail, promo=None, client=None):
    global API_ENABLED
    if API_ENABLED and client:
        prompt = (
            "สกัดรายละเอียดสินค้าเป็น JSON ภาษาไทยสั้นๆ\n"
            "กฎ:\n"
            "1. ใช้คำย่อชื่อสินค้า/แบรนด์ให้สั้นและกระชับ\n"
            "2. โครงสร้าง JSON object:\n"
            "   - 'ประเภท': ประเภทสินค้าสั้นๆ (เช่น นมโปรตีน, ทีวี)\n"
            "   - 'จุดเด่น': จุดเด่นหลัก 1-2 ข้อสั้นๆ\n"
            "   - 'ราคา': ตัวเลขราคาเท่านั้น (เช่น 399, 189) หรือ \"\" ถ้าไม่มี\n"
            "   - ฟิลด์ย่อยจำเป็นถ้ามี: 'รส', 'สี', 'ขนาด', 'รุ่น'\n"
            "3. ห้ามมีคำเกริ่น ห้ามใช้ markdown ```json ตอบกลับเฉพาะ JSON ดิบเท่านั้น\n\n"
            f"สินค้า:\n{detail}\n"
            f"โปร:\n{promo or ''}"
        )
        for model_idx, model in enumerate(TEXT_MODELS):
            if model_idx > 0:
                time.sleep(2)
            try:
                resp = client.models.generate_content(model=model, contents=prompt)
                res_text = resp.text.strip()
                res_text = re.sub(r'^```(?:json)?\s*', '', res_text)
                res_text = re.sub(r'\s*```$', '', res_text)
                import json
                parsed = json.loads(res_text)
                if isinstance(parsed, dict) and "ประเภท" in parsed:
                    return parsed
            except Exception as e:
                print(f"[{model}] parse_detail_to_json failed: {e}")
        print("[Warning] parse_detail_to_json failed on all models. Falling back to local parser.")
        
    return local_parse_detail_to_json(detail, promo)

def generate_local_fallback_caption(product_json, selected_persona, selected_hook, selected_style, path_norm, is_x, selected_role, in_post_body=False):
    prod_type = product_json.get("ประเภท", "ของกินของใช้")
    highlights = product_json.get("จุดเด่น", "คุณภาพดี ใช้งานสะดวก")
    price = format_thai_price(product_json.get("ราคา", ""))
    
    if "somtam" in path_norm:
        ending = "ค่ะ"
        closing = "แปะพิกัดลิงก์ร้านค้าไว้ในโพสต์แล้วนะคะ 👇" if in_post_body else "ดูลิ้งในคอมเมนต์แรกเลยนะคะ 👇"
    elif "chowchow" in path_norm:
        ending = "ฮะ"
        closing = "แปะพิกัดลิงก์ร้านค้าไว้ในโพสต์แล้วฮะ 👇" if in_post_body else "กดลิ้งในคอมเมนต์แรกได้เลยฮะ 👇"
    elif "x-bot" in path_norm:
        ending = "ครับ"
        closing = ""
    else:
        ending = "ครับ"
        closing = "แปะพิกัดลิงก์ร้านค้าไว้ในโพสต์แล้วครับ 👇" if in_post_body else "ดูลิ้งในคอมเมนต์แรกเลยครับ 👇"
        
    price_str = f" ราคาแค่ {price} บาท" if price else ""
    
    role_name = selected_role["name"]
    role_phrase = get_role_context(role_name, prod_type)
    style_name = selected_style["name"]
    if style_name == "มุกตลก":
        body = f"{selected_hook} {prod_type}ตัวนี้เลย{ending} {role_phrase} ตอนแรกกังวลว่าจะไม่โอเค แต่พอลองแล้วชอบเลย รอดตายแล้วเรา 555 จุดเด่นคือ {highlights}{price_str}"
    elif style_name == "ประสบการณ์ส่วนตัว" or "PAS" in style_name:
        body = f"{selected_hook} {role_phrase} ลองซื้อ {prod_type} ตัวนี้มาใช้ได้สักพักแล้ว{ending} รู้สึกสะดวกสบายขึ้นเยอะ โดยเฉพาะเรื่อง {highlights}{price_str} ดีจริง"
    elif style_name == "คำถาม":
        body = f"{selected_hook} {role_phrase} มีใครเคยลอง {prod_type} ตัวนี้รึยัง{ending} ส่วนตัวประทับใจจุดเด่น {highlights}{price_str} คิดว่าไงกันบ้าง"
    elif style_name == "การเปรียบเทียบ" or "BAB" in style_name:
        body = f"{selected_hook} {role_phrase} เทียบกับ {prod_type} แบบเดิมๆ ที่เคยใช้ ตัวนี้คือจุดเด่น {highlights}{price_str} ดีกว่าเห็นๆ เลย{ending}"
    else:
        body = f"{selected_hook} {role_phrase} แวะมาแชร์ {prod_type} ดีๆ{ending} ตัวนี้มีจุดเด่นคือ {highlights}{price_str} เผื่อใครกำลังสนใจอยู่"
        
    if is_x:
        return body[:200]
    else:
        return f"{body}\n\n{closing}"

def generate_caption(product_json, selected_persona, selected_hook, selected_style, shopee, lazada, promo, selected_role, in_post_body=False):
    global API_ENABLED
    import json
    
    path_norm = __file__.replace("\\", "/").lower()
    is_x = "x-bot" in path_norm
    
    if "somtam" in path_norm:
        gender_rule = "ใช้คำลงท้ายว่า 'ค่ะ' หรือ 'คะ' และสรรพนามแทนตัวว่า 'หนู' หรือ 'เรา' เท่านั้น"
        closing = "แปะพิกัดลิงก์ร้านค้าไว้ในโพสต์แล้วนะคะ 👇" if in_post_body else "ดูลิ้งในคอมเมนต์แรกเลยนะคะ 👇"
    elif "chowchow" in path_norm:
        gender_rule = "ใช้คำลงท้ายว่า 'ฮะ' หรือ 'โฮ่ง' และสรรพนามแทนตัวว่า 'น้องตูบ' หรือ 'ผม' เท่านั้น และมีกลิ่นอายความซนแบบน้องหมา"
        closing = "แปะพิกัดลิงก์ร้านค้าไว้ในโพสต์แล้วฮะ 👇" if in_post_body else "กดลิ้งในคอมเมนต์แรกได้เลยฮะ 👇"
    elif "x-bot" in path_norm:
        gender_rule = "ใช้คำลงท้าย 'ครับ' สรรพนามแทนตัว 'ผม'"
        closing = ""
    else:
        gender_rule = "ใช้คำลงท้ายว่า 'ครับ' และสรรพนามแทนตัวว่า 'ผม' หรือ 'พี่' เท่านั้น"
        closing = "แปะพิกัดลิงก์ร้านค้าไว้ในโพสต์แล้วครับ 👇" if in_post_body else "ดูลิ้งในคอมเมนต์แรกเลยครับ 👇"

    caption = None
    
    active_client = globals().get("client")
    if API_ENABLED and active_client:
        prompt = (
            "สวมบทบาทรีวิวแบบคนธรรมดา เป็นกันเอง\n"
            f"Persona: {selected_persona['desc']}\n"
            f"สไตล์การเขียน: {selected_style['desc']}\n\n"
            f"เริ่มต้นประโยคแรกด้วย Hook นี้เป๊ะๆ ห้ามดัดแปลง: \"{selected_hook}\"\n\n"
            f"ข้อมูลสินค้า JSON:\n{json.dumps(product_json, ensure_ascii=False)}\n\n"
            f"กฎเหล็ก:\n- ต้องระบุราคาสินค้าจาก JSON เสมอ\n"
            f"- ห้ามเริ่มโพสต์ด้วยชื่อแบรนด์/สินค้า\n"
            f"- ห้ามใช้คำโฆษณาจำพวก คุ้มมาก, คุ้มสุดๆ, คุ้มค่า, คุ้ม, ดีงาม, ห้ามพลาด, ของดี, ดีจริง, แนะนำเลย\n"
            f"- ห้ามแต่งข้อมูลที่ไม่มีใน JSON เด็ดขาด: ห้ามอ้างของใกล้หมด/ล็อตสุดท้าย/จำนวนจำกัด, ห้ามเคลมสรรพคุณสุขภาพ ลดน้ำหนัก คุมอาหาร แคลอรี่ โปรตีน ถ้า JSON ไม่ได้ระบุ, ห้ามเดาวิธีใช้ที่ผิดธรรมชาติของสินค้าชนิดนั้น\n"
            f"- เขียนถึงสินค้าให้ตรงกับสิ่งที่มันเป็นจริง ถ้าไม่แน่ใจว่าสินค้าคืออะไร ให้เล่ากลางๆ ตามชื่อสินค้า อย่ามโนรายละเอียดเพิ่ม\n"
            f"- ถ้าบทบาท/Persona ไม่เข้ากับสินค้าชนิดนี้ ห้ามฝืนโยงสินค้าเข้ากับบทบาท ให้เล่าแบบคนทั่วไปที่บังเอิญเจอของน่าสนใจแทน\n"
            f"- โทนเสียง: {gender_rule}\n"
            "- เน้นทำ Social SEO: ค้นหาและใส่คำค้นหา (keywords) ยอดฮิตที่คนมักจะพิมพ์ค้นหาเกี่ยวกับสินค้านี้ลงในเนื้อหาแบบเนียนๆ เป็นธรรมชาติ\n"
            "- กระตุ้นการแชร์และเซฟโพสต์เก็บไว้ (เช่น 'เซฟเก็บไว้ดูตอนช้อป', 'แชร์ต่อให้เพื่อนที่เจอปัญหานี้') ตามความเหมาะสมกับบริบท\n\n"
        )
        
        if is_x:
            prompt += (
                "รูปแบบโพสต์สำหรับ X (Twitter):\n"
                "1. ความยาวสั้นกระชับ รวมกันไม่เกิน 150-200 ตัวอักษร\n"
                "2. เล่าความเห็นส่วนตัวสั้นๆ สอดรับกับ Persona และ Style\n"
                "3. ลงท้ายด้วยสรรพนามธรรมชาติ ไม่ต้องมีข้อความปิดท้ายอื่น\n"
                "ตอบกลับเฉพาะเนื้อความโพสต์เท่านั้น"
            )
        else:
            prompt += (
                f"รูปแบบโพสต์สำหรับ Facebook:\n"
                f"1. ความยาวประมาณ 2-5 ประโยค\n"
                f"2. เล่าเรื่องราวหรือให้ความเห็นส่วนตัวสั้นๆ สอดรับกับ Persona และ Style\n"
                f"3. ปิดท้ายด้วยประโยคนี้เป๊ะๆ: \"{closing}\"\n"
                f"ตอบกลับเฉพาะเนื้อความโพสต์เท่านั้น ไม่ต้องมีข้อความนำ/อธิบายใดๆ"
            )
            
        for model_idx, model in enumerate(TEXT_MODELS):
            if model_idx > 0:
                time.sleep(2)
            try:
                resp = active_client.models.generate_content(model=model, contents=prompt)
                caption_text = resp.text.strip()
                if caption_text:
                    lines = caption_text.splitlines()
                    while lines and (
                        re.search(r'^(ได้เลย|นี่คือ|แน่นอน|โพสต์รีวิว|ครับ|ค่ะ|---)', lines[0].strip(), re.IGNORECASE)
                        or lines[0].strip() in ("", "---")
                    ):
                        lines.pop(0)
                    caption = "\n".join(lines).strip()
                    break
            except Exception as e:
                err_msg = str(e)
                print(f"[{model}] caption generation failed: {err_msg[:80]}")
                if any(x in err_msg.lower() for x in ["api key", "invalid_argument", "api_key", "timeout", "timed out", "deadline exceeded", "connection", "connect", "unreachable"]):
                    print("API key or connection issue detected. Disabling API calls.")
                    API_ENABLED = False
                    break
                    
        if not caption and API_ENABLED:
            print("[Warning] Caption generation failed on all models. Disabling API calls.")
            API_ENABLED = False

    if not caption:
        print("[Warning] Falling back to local heuristic caption.")
        caption = generate_local_fallback_caption(product_json, selected_persona, selected_hook, selected_style, path_norm, is_x, selected_role, in_post_body=in_post_body)
        
    promo_line = f"\n🔥 โปรโมชั่น: {promo}" if promo else ""
    if is_x:
        promo_line = f" 🔥 {promo}" if promo else ""
        
    if promo and promo_line not in caption:
        caption += promo_line
        
    return caption

def _post_one_comment(post_id, text):
    try:
        resp = requests.post(
            f"https://graph.facebook.com/v21.0/{post_id}/comments",
            data={"access_token": PAGE_ACCESS_TOKEN, "message": text},
            timeout=30
        )
        result = resp.json()
        if "id" in result:
            print(f"Comment posted: {result['id']}")
        else:
            print(f"Comment failed: {result}")
    except Exception as e:
        print(f"Comment error: {e}")

def post_link_comment(post_id, shopee, lazada, promo):
    """โพส comment ลิ้งใต้โพส แยก Shopee / Lazada คนละคอมเม้น"""
    promo_line = f"\n🔥 โปร: {promo}" if promo else ""
    if shopee and "xxx" not in shopee:
        _post_one_comment(post_id, f"👉 ซื้อได้ที่ Shopee → {shopee}{promo_line}")
    if lazada and "xxx" not in lazada:
        _post_one_comment(post_id, f"🛍️ หรือสั่งทาง Lazada → {lazada}")

def post_to_page(img_path, caption, shopee=None, lazada=None, promo=None, scheduled_timestamp=None):
    print("Posting to Facebook Page...")
    from affiliate_utils import get_next_scheduled_time

    if scheduled_timestamp is not None:
        scheduled_time = scheduled_timestamp
        print(f"Using explicit scheduled_timestamp: {scheduled_time}")
    else:
        slots = ["08:00", "10:00", "12:30", "15:00", "18:00"]
        scheduled_time = get_next_scheduled_time(slots)
    
    if scheduled_time:
        comment_texts = []
        promo_line = f"\n🔥 โปร: {promo}" if promo else ""
        if shopee and "xxx" not in shopee:
            comment_texts.append(f"👉 ซื้อได้ที่ Shopee → {shopee}{promo_line}")
        if lazada and "xxx" not in lazada:
            comment_texts.append(f"🛍️ หรือสั่งทาง Lazada → {lazada}")
            
        if comment_texts:
            caption += "\n\n📌 ชี้เป้าของดีน่าสนใจ:\n" + "\n".join(comment_texts)
            
        print(f"Scheduling to Facebook for timestamp {scheduled_time}...")
        with open(img_path, "rb") as f:
            resp = requests.post(
                f"https://graph.facebook.com/v25.0/{PAGE_ID}/photos",
                data={
                    "access_token": PAGE_ACCESS_TOKEN,
                    "message": caption,
                    "published": "false",
                    "unpublished_content_type": "SCHEDULED",
                    "scheduled_publish_time": scheduled_time
                },
                files={"source": ("review.png", f, "image/png")},
                timeout=60
            )
        result = resp.json()
        if "id" in result:
            photo_id = result.get("post_id") or result["id"]
            print(f"Scheduled successfully! Photo ID: {photo_id}")
            return photo_id, True
        else:
            print(f"FB Error: {result}")
            raise SystemExit(1)

    with open(img_path, "rb") as f:
        resp = requests.post(
            f"https://graph.facebook.com/v25.0/{PAGE_ID}/photos",
            data={"access_token": PAGE_ACCESS_TOKEN, "message": caption, "published": "true"},
            files={"source": ("review.png", f, "image/png")},
            timeout=60
        )
    result = resp.json()
    if "id" in result:
        post_id = result.get("post_id") or result["id"]
        print(f"Page Posted! ID: {post_id}")
        print(f"https://www.facebook.com/{post_id}")
        return post_id, False
    else:
        print(f"FB Error: {result}")
def extract_badge_text(promo):
    if not promo:
        return None
    pct_match = re.search(r'(ลด\s*\d+\s*%)|(\d+\s*%\s*OFF)', promo, re.IGNORECASE)
    if pct_match:
        val = pct_match.group(0)
        val = re.sub(r'\s+', ' ', val)
        return val
    price_match = re.search(r'฿\s*\d+', promo)
    if price_match:
        return price_match.group(0).replace(" ", "")
    return None

if __name__ == "__main__":
    import argparse, time as _time, json
    from datetime import datetime, timezone, timedelta
    init_db()
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="Run without posting or marking as done")
    args = parser.parse_args()

    IMMEDIATE = os.environ.get("IMMEDIATE", "false").lower() == "true"

    # คำนวณ 5 scheduled timestamps ล่วงหน้า (Bangkok UTC+7)
    # IMMEDIATE=true (workflow_dispatch) -> post now, no scheduling
    BKK = timezone(timedelta(hours=7))
    now_bkk = datetime.now(BKK)
    DAILY_SLOTS = ["18:20"]
    slot_timestamps = []
    for idx, slot in enumerate(DAILY_SLOTS):
        if IMMEDIATE:
            dt = now_bkk + timedelta(minutes=15 + idx * 120)
            slot_timestamps.append(int(dt.timestamp()))
        else:
            h, m = map(int, slot.split(":"))
            dt = datetime(now_bkk.year, now_bkk.month, now_bkk.day, h, m, tzinfo=BKK)
            if dt <= now_bkk + timedelta(minutes=10):
                dt += timedelta(days=1)
            slot_timestamps.append(int(dt.timestamp()))

    # โหลด state เพียงครั้งเดียวก่อนเข้าลูป เพื่อให้จำค่าการสุ่มข้ามโพสต์ได้ถูกต้อง
    import openpyxl
    wb_init = openpyxl.load_workbook(EXCEL_PATH)
    state = load_state(wb_init)
    wb_init.close()

    posted_this_run = set()   # dedup shopee URL ภายใน run เดียวกัน
    success_count = 0

    for i in range(len(DAILY_SLOTS)):
        API_ENABLED = True   # reset ทุก iteration ให้ Gemini ลองใหม่

        print(f"\n===== Post {i+1}/{len(DAILY_SLOTS)} =====")
        scheduled_ts = slot_timestamps[i] if i < len(slot_timestamps) else None

        product, wb, ws = load_next_product()
        affiliate_mode = False
        if not product:
            print("review_products.xlsx หมดแล้ว — ลอง fallback จาก AFFILIATE_DIR")
            posted_urls = get_posted_urls(ws)
            posted_urls |= posted_this_run   # รวม URL ที่โพสใน run นี้ด้วย
            product = load_affiliate_product(posted_urls)
            affiliate_mode = True
            if not product:
                print("ไม่มีสินค้าเหลือ หยุด")
                sys.exit(1)

        # ป้องกันซ้ำใน run เดียวกัน
        if product.get("shopee") in posted_this_run:
            print(f"[Skip] ซ้ำใน run นี้: {str(product.get('shopee',''))[:60]}")
            continue

        print(f"Product: {product['detail'][:60]}...")

        promo_clean = clean_promo(product["promo"])
        product_json = parse_detail_to_json(product["detail"], promo_clean, client)
        print(f"[System Log] Converted Product JSON:\n{json.dumps(product_json, ensure_ascii=False, indent=2)}")

        # Loop to generate unique caption with similarity checks
        max_attempts = 10
        caption = None
        chosen_p = None
        chosen_h = None
        chosen_s = None
        chosen_r = None

        for attempt in range(max_attempts):
            selected_persona = choose_persona(state)
            selected_hook = choose_hook(state)
            selected_style = choose_style(state)
            selected_role = choose_role(state)

            candidate_caption = generate_caption(
                product_json, selected_persona, selected_hook, selected_style,
                product["shopee"], product["lazada"], promo_clean,
                selected_role
            )

            # Dual similarity check: SQLite (vector + fallback Jaccard) and local Excel recent_captions
            is_similar = is_duplicate(candidate_caption, client, threshold=0.82, jaccard_threshold=0.60)
            if not is_similar:
                recent_caps = state.get("recent_captions", [])[-20:]
                for old_cap in recent_caps:
                    score = get_caption_similarity(candidate_caption, old_cap)
                    if score > 0.60:
                        is_similar = True
                        print(f"[Attempt {attempt+1}] Caption rejected by Excel memory! Similarity too high ({score:.2f} > 0.60)")
                        break

            if not is_similar:
                caption = candidate_caption
                chosen_p = selected_persona
                chosen_h = selected_hook
                chosen_s = selected_style
                chosen_r = selected_role
                print(f"[System Log] Caption approved on attempt {attempt+1}!")
                break

        if not caption:
            print("[Warning] Could not generate unique caption in 10 attempts. Using the last candidate.")
            caption = candidate_caption
            chosen_p = selected_persona
            chosen_h = selected_hook
            chosen_s = selected_style
            chosen_r = selected_role

        print(f"[System Log] Chosen Role: {chosen_r['name']} ({chosen_r['id']})")
        print(f"[System Log] Chosen Persona: {chosen_p['name']} ({chosen_p['id']})")
        print(f"[System Log] Chosen Hook: {chosen_h}")
        print(f"[System Log] Chosen Style: {chosen_s['name']}")

        # Image generation using overlay
        highlights = product_json.get("จุดเด่น", "")
        line1, line2 = generate_hook(product["detail"], highlights)
        if "segment_thai_text" in globals():
            line1 = segment_thai_text(line1, client)
            line2 = segment_thai_text(line2, client)
        print(f"Hook: {line1} | {line2}")

        product_img = download_image(product["image_url"])
        line1_clean = clean_overlay_text(line1)
        line2_clean = clean_overlay_text(line2)
        if not line1_clean:
            line1_clean = "สินค้าแนะนำ"
        review_img = None
        try:
            from card.render_card import render_review_card
            card_out = product_img.rsplit(".", 1)[0] + "_card.png"
            review_img = render_review_card(product_img, line1_clean, line2_clean, card_out,
                                            price=product_json.get("ราคา"))
            os.unlink(product_img)
            print(f"Card render done: {review_img}")
        except Exception as card_err:
            print(f"[Warning] Card render failed ({card_err}), falling back to PIL overlay")
        if not review_img:
            try:
                badge_text = extract_badge_text(product.get("promo"))
                review_img = add_overlay(
                    product_img, line1_clean, line2_clean, ACCENT_COLOR,
                    font_name="Itim-Regular.ttf",
                    badge_text=badge_text,
                    watermark="พริก 10 เม็ด"
                )
                os.unlink(product_img)
                print(f"Overlay done: {review_img}")
            except Exception as overlay_err:
                print(f"Overlay failed, using original: {overlay_err}")
                review_img = product_img
        print(f"Caption:\n{caption}\n")

        # Save selection details
        state["last_personas"].append(chosen_p["id"])
        state["last_hooks"].append(chosen_h)
        state["last_styles"].append(chosen_s["name"])
        state["last_roles"].append(chosen_r["id"])
        state["recent_captions"].append(caption)

        if args.dry_run:
            print(f"[Dry run] img={review_img} | shopee={product['shopee']}")
        else:
            post_id, was_scheduled = post_to_page(
                review_img, caption,
                product["shopee"], product["lazada"], promo_clean, 
                scheduled_timestamp=scheduled_ts
            )
            if os.path.exists(review_img):
                os.unlink(review_img)
            
            # Save to SQLite memory
            slot_name = DAILY_SLOTS[i] if i < len(DAILY_SLOTS) else "unknown"
            save_post_to_db(caption, product["detail"], slot_name, client)

            if not was_scheduled:
                post_link_comment(post_id, product["shopee"], product["lazada"], promo_clean)
            if not affiliate_mode:
                mark_posted(wb, ws, product["row"], state=state)
            else:
                append_posted_fallback(wb, ws, product, state=state)
            posted_this_run.add(product["shopee"])
            success_count += 1

        if i < 4:
            _time.sleep(5)

    print(f"\nDone: {success_count}/5 posts completed")

