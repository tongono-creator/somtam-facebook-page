# affiliate_utils.py — อ่าน product/food links จาก affiliate_products.xlsx
import os, re, random
from datetime import datetime, timezone, timedelta

WEBSITE_URL = "https://shopee-ranking.vercel.app/"
EXCEL_PATH  = os.path.join(os.path.dirname(__file__), "affiliate_products.xlsx")

BKK = timezone(timedelta(hours=7))

def _now_bkk():
    return datetime.now(BKK)

def _rotate(items, extra_salt=0):
    """สุ่มเลือกจาก list — ไม่ซ้ำกันระหว่างโพส"""
    if not items:
        return None
    return random.choice(items)

def parse_thai_date(date_str):
    """
    พยายามแปลงวันที่ไทยเป็น date object เช่น '25 พ.ค. 69', '2026-05-29'
    ปี พ.ศ. 2569 / 69 จะถูกแปลงเป็น ค.ศ. 2026
    """
    if not date_str:
        return None
    date_str = str(date_str).strip().lower()
    
    # 1. Check YYYY-MM-DD standard format FIRST before splitting!
    match = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})', date_str)
    if match:
        try:
            return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3))).date()
        except ValueError:
            pass

    # 2. range splitting: split by any hyphen or slash representing range and take the end date
    parts = re.split(r'[-–ถึง]', date_str)
    if len(parts) > 1:
        date_str = parts[-1].strip()

    # Check YYYY-MM-DD again in case it was a range of ISO dates
    match = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})', date_str)
    if match:
        try:
            return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3))).date()
        except ValueError:
            pass

    # Check DD/MM/YYYY
    match_slash = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{2,4})', date_str)
    if match_slash:
        try:
            d = int(match_slash.group(1))
            m = int(match_slash.group(2))
            y = int(match_slash.group(3))
            if y > 2500: y -= 543
            elif y > 2400: y -= 543
            elif y < 100:
                if y >= 50:
                    y = y + 1957
                else:
                    y = y + 2000
            return datetime(y, m, d).date()
        except ValueError:
            pass
        
    # Thai months dictionary
    thai_months = {
        'ม.ค.': 1, 'มกราคม': 1,
        'ก.พ.': 2, 'กุมภาพันธ์': 2,
        'มี.ค.': 3, 'มีนาคม': 3,
        'เม.ย.': 4, 'เมษายน': 4,
        'พ.ค.': 5, 'พฤษภาคม': 5,
        'มิ.ย.': 6, 'มิถุนายน': 6,
        'ก.ค.': 7, 'กรกฎาคม': 7,
        'ส.ค.': 8, 'สิงหาคม': 8,
        'ก.ย.': 9, 'กันยายน': 9,
        'ต.ค.': 10, 'ตุลาคม': 10,
        'พ.ย.': 11, 'พฤศจิกายน': 11,
        'ธ.ค.': 12, 'ธันวาคม': 12
    }
    
    # Match Thai date format with optional spaces
    match_thai = re.search(r'(\d{1,2})\s*([ก-๙\.]+)\s*(\d{2,4})', date_str)
    if match_thai:
        try:
            d = int(match_thai.group(1))
            m_name = match_thai.group(2)
            y = int(match_thai.group(3))
            
            m = None
            for k, v in thai_months.items():
                if k in m_name or m_name in k:
                    m = v
                    break
                    
            if m and y:
                if y > 2500: y -= 543
                elif y > 2400: y -= 543
                elif y < 100:
                    if y >= 50:
                        y = y + 1957
                    else:
                        y = y + 2000
                return datetime(y, m, d).date()
        except ValueError:
            pass
            
    return None

# ─── อ่าน Excel ──────────────────────────────────────────────────
def _load_excel():
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb.active
        products, food_entries, promos = [], [], []
        today = _now_bkk().date()

        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 5:
                continue
            no, name, shopee, lazada, active = row[0], row[1], row[2], row[3], row[4]
            desc     = row[5] if len(row) > 5 else ""
            food_raw = row[6] if len(row) > 6 else None
            # column H=promo_message, I=promo_url, J=promo_active, K=expiry_date, L=picture_url
            promo_msg    = row[7]  if len(row) > 7  else None
            promo_url    = row[8]  if len(row) > 8  else None
            promo_active = row[9]  if len(row) > 9  else None
            expiry_raw   = row[10] if len(row) > 10 else None
            picture_url  = row[11] if len(row) > 11 else None

            # product rows (ต้องมี active=yes)
            if str(active).strip().lower() == "yes":
                products.append({
                    "name":   str(name or "").strip(),
                    "shopee": str(shopee or "").strip(),
                    "lazada": str(lazada or "").strip(),
                    "desc":   str(desc or "").strip(),
                    "price":  "",
                    "category": "main",
                    "image":  "",
                })

            # food links — เก็บทุก row ที่มีค่า column G
            if food_raw:
                food_entries.append(str(food_raw).strip())

            # promo — column H-L, active=yes + ยังไม่หมดอายุ
            promo_ok = str(promo_active or "").strip().lower() == "yes"
            if promo_ok and expiry_raw:
                try:
                    if hasattr(expiry_raw, "date"):
                        exp = expiry_raw.date()
                    else:
                        exp = parse_thai_date(expiry_raw)
                    if exp:
                        if today > exp:
                            promo_ok = False
                            print(f"Promo expired ({exp}): {str(promo_msg or '')[:40]}")
                    else:
                        promo_ok = False
                        print(f"Promo expiry date unparseable ('{expiry_raw}'), defaulting to expired: {str(promo_msg or '')[:40]}")
                except Exception as parse_err:
                    promo_ok = False
                    print(f"Promo expiry parse error ('{expiry_raw}'), defaulting to expired: {parse_err}")

            if promo_msg and promo_ok:
                promos.append({
                    "message":     str(promo_msg).strip(),
                    "url":         str(promo_url or "").strip(),
                    "picture_url": str(picture_url or "").strip(),
                })

        # Load extra products from affiliate_data/*.xlsx
        extra_dir = os.path.join(os.path.dirname(__file__), "affiliate_data")
        if os.path.exists(extra_dir):
            for file_name in os.listdir(extra_dir):
                if file_name.endswith(".xlsx") and not file_name.startswith("~$"):
                    try:
                        extra_path = os.path.join(extra_dir, file_name)
                        extra_wb = openpyxl.load_workbook(extra_path, data_only=True)
                        extra_ws = extra_wb.active
                        for row in extra_ws.iter_rows(min_row=2, values_only=True):
                            if len(row) < 10:
                                continue
                            name = row[1]
                            image = row[2] if len(row) > 2 else ""
                            price = row[3]
                            shopee = row[9]
                            lazada = row[10] if len(row) > 10 else None

                            if name and shopee:
                                shopee_str = str(shopee).strip()
                                if not shopee_str.startswith("http"):
                                    shopee_str = ""
                                if shopee_str and "xxx" not in shopee_str:
                                    lazada_str = str(lazada).strip() if lazada else ""
                                    if not lazada_str.startswith("http"):
                                        lazada_str = ""
                                    p_str = ""
                                    if price:
                                        p_str = str(price).strip()
                                        if p_str.endswith(".00"):
                                            p_str = p_str[:-3]
                                    desc = str(name).strip()

                                    products.append({
                                        "name":   str(name).strip(),
                                        "shopee": shopee_str,
                                        "lazada": lazada_str,
                                        "desc":   desc,
                                        "price":  p_str,
                                        "category": file_name,
                                        "image":  str(image).strip() if image else "",
                                    })
                    except Exception as extra_err:
                        print(f"Error loading extra excel {file_name}: {extra_err}")

        # Load from review_products.xlsx too (especially for review post comment lookups)
        review_path = os.path.join(os.path.dirname(__file__), "review_products.xlsx")
        if os.path.exists(review_path):
            try:
                review_wb = openpyxl.load_workbook(review_path, data_only=True)
                review_ws = review_wb.active
                for row in review_ws.iter_rows(min_row=2, values_only=True):
                    if len(row) < 4:
                        continue
                    detail = row[1]
                    shopee = row[2]
                    lazada = row[3]
                    
                    if detail and shopee:
                        shopee_str = str(shopee).strip()
                        if shopee_str.startswith("http") and "xxx" not in shopee_str:
                            lazada_str = str(lazada).strip() if lazada else ""
                            if not lazada_str.startswith("http"):
                                lazada_str = ""
                            
                            detail_str = str(detail).strip()
                            name = detail_str.split("\n")[0][:60]
                            products.append({
                                "name":   name,
                                "shopee": shopee_str,
                                "lazada": lazada_str,
                                "desc":   detail_str,
                                "image":  "",
                            })
            except Exception as review_err:
                print(f"Error loading review excel in affiliate_utils: {review_err}")

        return products, food_entries, promos
    except Exception as e:
        print(f"Excel read error: {e}")
        return [], [], []

def _parse_food(raw):
    """แยก shop_name + url จาก 'ลองเข้ามาดู ShopName ที่ Shopee! https://...'"""
    url_match = re.search(r'https://\S+', raw)
    url = url_match.group(0) if url_match else raw
    name_match = re.search(r'ลองเข้ามาดู\s+(.+?)\s+ที่ Shopee', raw)
    name = name_match.group(1).strip() if name_match else "ร้านอาหาร"
    return name, url

# ─── Website comment variations ──────────────────────────────────
WEBSITE_VARS = [
    f"📍 เผื่อใครหาอันดับสินค้าประเภทต่างๆ บน Shopee เข้าไปดูเพิ่มเติมได้ที่นี่เลยครับ → {WEBSITE_URL}",
    f"📊 เช็คอันดับและสถิติของใช้ยอดฮิตบน Shopee ก่อนช้อปได้ที่นี่เลย → {WEBSITE_URL}",
    f"🏆 สถิติของใช้ยอดฮิตวันนี้บน Shopee เข้าไปเช็คข้อมูลกันได้เลยครับ → {WEBSITE_URL}",
    f"💡 ใครอยากได้ลิสต์ของใช้แต่ละหมวดหมู่เทียบกัน ลองดูข้อมูลที่นี่ได้ครับ → {WEBSITE_URL}",
    f"🛒 ข้อมูลของใช้แต่ละหมวดหมู่บน Shopee สรุปไว้ให้ดูเปรียบเทียบตรงนี้ครับ → {WEBSITE_URL}",
]

# ─── Food comment variations ─────────────────────────────────────
# Templates will be dynamically selected and structured in get_food_comment()

# ─── Shopee / Lazada fallback comment templates (3-Step: Hook -> Story -> CTA) ───────────────────────────
SHOPEE_INTROS = [
    "📍 เผื่อใครถามพิกัดของ {name} ราคา {price} บาท วางลิงก์ Shopee ไว้ให้ตรงนี้นะครับ → {url}",
    "💬 มีคนถามถึง {name} ราคา {price} บาท บ่อยๆ วางพิกัด Shopee ไว้ให้ทางนี้นะครับ → {url}",
    "💡 {name} ราคา {price} บาท ตัวที่เล่าไป ใครสนใจพิกัด Shopee ดูได้ตรงนี้ครับ → {url}",
]

LAZADA_INTROS = [
    "📍 เผื่อใครถามพิกัดของ {name} ราคา {price} บาท วางลิงก์ Lazada ไว้ให้ตรงนี้นะครับ → {url}",
    "💬 มีคนถามถึง {name} ราคา {price} บาท บ่อยๆ วางพิกัด Lazada ไว้ให้ทางนี้นะครับ → {url}",
    "💡 {name} ราคา {price} บาท ตัวที่เล่าไป ใครสนใจพิกัด Lazada ดูได้ตรงนี้ครับ → {url}",
]

PRODUCT_HOOKS = [
    "ชี้เป้าของใช้",
    "พิกัดของชิ้นนี้",
    "รายละเอียดสินค้า",
]

PROMO_INTROS = [
    "📍 ส่วนลดพิเศษสำหรับ {name} ราคาพิเศษ {price} บาท ดูพิกัดตรงนี้ได้เลยครับ → {url}",
    "⚡ ราคาพิเศษช่วงนี้กับ {name} ดูรายละเอียดเพิ่มเติมในลิงก์ได้เลยครับ → {url}",
    "🎯 พิกัดราคาพิเศษของ {name} สนใจสั่งซื้อดูตรงนี้ได้เลยครับ → {url}",
]

def get_website_with_product_comment():
    """ดึงข้อความแนะนำเว็บ shopee-ranking และสุ่มต่อท้ายด้วยสินค้าแนะนำที่กำลังแอคทีฟ 1 ชิ้นเพื่อเพิ่มโอกาสขาย"""
    web_base = _rotate(WEBSITE_VARS)
    if not web_base:
        web_base = f"📍 เผื่อใครหาอันดับสินค้าประเภทต่างๆ บน Shopee เข้าไปดูเพิ่มเติมได้ที่นี่เลยครับ → {WEBSITE_URL}"
        
    try:
        products, _, _ = _load_excel()
        active = [p for p in products if p.get("shopee") and "xxx" not in p.get("shopee")]
        if active:
            p = random.choice(active)
            shopee_url = p.get("shopee", "").strip()
            lazada_url = p.get("lazada", "").strip()
            if "xxx" in shopee_url: shopee_url = ""
            if "xxx" in lazada_url: lazada_url = ""
            
            links = []
            if shopee_url:
                links.append(f"🧡 Shopee: {shopee_url}")
            if lazada_url:
                links.append(f"💙 Lazada: {lazada_url}")
                
            if links:
                link_section = "\n".join(links)
                web_base += f"\n\n📌 ชี้เป้าของดีแนะนำวันนี้ - {p.get('name', 'สินค้าแนะนำ')}:\n{link_section}"
    except Exception as e:
        print(f"Error appending product to website comment: {e}")
        
    return web_base

# ─── Public API ───────────────────────────────────────────────────
def get_standard_comments():
    """comment เว็บ shopee-ranking (หมุน variations)"""
    return [get_website_with_product_comment()]

def get_persona():
    """ตรวจจับว่ากำลังรันอยู่ในโฟลเดอร์ของเพจไหนเพื่อเลือกบุคลิกภาพที่ถูกต้อง"""
    path = os.path.abspath(__file__).replace("\\", "/")
    if "chowchow" in path:
        return "chowchow"
    elif "somtam" in path:
        return "somtam"
    elif "kram" in path:
        return "kram"
    elif "x-bot" in path:
        return "xbot"
    else:
        return "rocket"

def generate_comment_with_ai(p, platform, persona, caption=None):
    """ใช้ Gemini-2.5-flash เขียนคอมเมนต์ 3 ขั้นตอนสไตล์แอดมินตามเพจ"""
    api_key = os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
    if not api_key:
        return None

    # กำหนดบุคลิกตามเพจ
    if persona == "chowchow":
        persona_inst = "น้องหมา Chow Chow เพศผู้ (ชื่อแอดมินน้องตูบ) เล่าเรื่องราวแสนซน ขี้อ้อน และตรงไปตรงมา มีคำลงท้ายภาษาตูบตัวผู้เสมอ เช่น ฮะ, ครับ หรือมีเสียงร้อง โฮ่ง บ้าง แทนตัวว่า ผม หรือ น้องตูบ"
    elif persona == "somtam":
        persona_inst = "ผู้หญิงที่เป็นกันเอง น่ารัก ตลก และตรงไปตรงมา (เพจชื่อพริก 10 เม็ด) มีคำลงท้ายภาษาผู้หญิงเสมอ เช่น ค่ะ/คะ แทนตัวว่า หนู หรือ เรา"
    else: # kram, xbot, rocket (ผู้ชาย)
        persona_inst = "ผู้ชาย สุภาพและเป็นกันเอง ลงท้ายด้วยครับ/ผม หรือพี่"

    name = p.get("name", "สินค้าแนะนำ")
    desc = p.get("desc", "").strip()

    prompt = (
        f"คุณคือแอดมินเพจโซเชียลมีเดียที่เป็น: {persona_inst}\n\n"
        f"ช่วยเขียนคอมเมนต์แนะนำสินค้าเพื่อโปรโมทลิงก์แอฟฟิลิเอต (Affiliate) บน {platform} โดยใช้เทคนิค 3 ขั้นตอนในการเขียน:\n"
        f"1. เปิดให้น่าสนใจ (Hook): ประโยคเปิดหัวสั้นๆ กระชับ โดยพยายามเชื่อมโยงหรือเปรียบเทียบเนื้อหาของโพสต์หลัก (ถ้ามี) เข้ากับสินค้าอย่างเนียนๆ ตลกขบขัน หรือเปรียบเปรยประเด็นชีวิตคนทำงาน/ศึกสงครามให้ดึงดูดใจ\n"
        f"2. เล่าให้เห็นภาพ (Vivid Storytelling): บรรยายสั้นๆ ให้คนเห็นภาพประโยชน์การใช้งานสินค้า\n"
        f"3. ปิดจบต้องบอกว่า 'ควรทำอะไร' (Call to Action): ชี้เป้าให้กดตะกร้าสั่งซื้อ บังคับมีคำลงท้ายตามบุคลิกภาพของคุณ\n\n"
    )
    if caption:
        prompt += f"ข้อความแคปชั่นโพสต์หลักเพื่อใช้ในการเชื่อมโยงมุก:\n\"\"\"\n{caption}\n\"\"\"\n\n"
        
    prompt += (
        f"รายละเอียดสินค้า:\n"
        f"- ชื่อสินค้า: {name}\n"
        f"- ข้อมูลเด่น: {desc}\n\n"
        f"กฎการร่างข้อความ:\n"
        f"- ห้ามใช้ bullet points, รายการตัวเลข, หรือทำเป็นข้อๆ และห้ามมีสัญลักษณ์รายการใดๆ (เช่น •, ▪️, -, 👉) ปรากฏในเนื้อหาเด็ดขาด ให้เขียนเป็นย่อหน้าเดียวต่อเนื่องเป็นธรรมชาติ\n"
        f"- ห้ามใส่ลิงก์เด็ดขาด (ลิงก์จะถูกนำไปต่อท้ายเอง)\n"
        f"- ความยาวรวมไม่เกิน 2-3 บรรทัด (สั้น กระชับ คล้ายสไตล์คุยกันใต้โพสต์)\n"
        f"- ห้ามใช้ markdown เช่น ตัวหนา ** หรืออัญประกาศ\n"
        f"- ตอบเฉพาะตัวข้อความภาษาไทยเท่านั้น"
    )

    try:
        from google import genai
        client = genai.Client(api_key=api_key, http_options={'timeout': 90.0})
        resp = client.models.generate_content(
            model="gemini-flash-latest",
            contents=prompt
        )
        text = resp.text.strip()
        # Clean potential echoes
        text = re.sub(r'^(คอมเมนต์|Comment|ข้อความ|คำอธิบาย)[:\-\s\.]+', '', text, flags=re.IGNORECASE).strip()
        text = text.strip('"\'“”‘’')
        return text
    except Exception as e:
        print(f"AI Comment generation failed: {e}")
        return None

def get_food_comment():
    """comment Shopee Food หมุนเวียนทุกร้าน ตามหลัก 3 ขั้นตอน"""
    _, food_entries, _ = _load_excel()
    raw = _rotate(food_entries, extra_salt=3)
    if not raw:
        return None
    name, url = _parse_food(raw)
    
    persona = get_persona()
    if persona == "chowchow":
        ending = "ฮะ โฮ่ง!"
    elif persona == "somtam":
        ending = "ค่ะ"
    else:
        ending = "ครับ"

    food_templates = [
        "🍜 หิวตอนดึกแบบนี้ต้องจัดแล้วครับกับ {name} กลิ่นหอมๆ รสชาติเข้มข้นถึงใจแน่นอน ตอนนี้มีโปรเด็ดลดราคาพิเศษอยู่นะ กดตะกร้าสั่งผ่าน Shopee Food ได้เลย{ending} → {url}",
        "🍱 แนะนำร้านอร่อยนี้เลย {name} รสเด็ด เมนูเด่น จัดเต็มคำ กินกี่ทีก็ไม่เบื่อ สนใจกดสั่งในตะกร้าด่วน มีโปรโมชั่นสุดพิเศษรออยู่นะ{ending} → {url}",
        "🔖 หิวแบบนี้จะพลาดได้ไงกับ {name} อาหารสดใหม่ รสชาติกลมกล่อมฟินทุกคำ กดสั่งทางนี้เลย มีโปรลดจุกๆ วันนี้เท่านั้นนะ{ending} → {url}",
        "🛵 เมนูอร่อยชวนน้ำลายสอเลยกับ {name} ร้อนๆ คุ้มค่าสมราคาพร้อมส่งทันที อย่ารอช้า กดตะกร้าสั่งได้เลย มีโปรส่งฟรีอยู่นะ{ending} → {url}",
    ]
    
    template = random.choice(food_templates)
    return template.format(name=name, url=url, ending=ending)

def get_promo_comment():
    """comment โปรโมชั่น — ใช้ message จาก Excel โดยตรง (รองรับ picture_url)"""
    _, _, promos = _load_excel()
    if not promos:
        return None
    p = random.choice(promos)
    msg = p["message"].strip()
    if not msg:
        print("Promo skipped: empty message")
        return None
    if p.get("url"):
        msg += f" → {p['url']}"
    pic = p.get("picture_url", "").strip()
    # ตรวจว่าเป็น URL จริง ไม่ใช่ชื่อตัวแปรหรือ placeholder
    if not pic.startswith("http"):
        pic = ""
    return {"message": msg, "picture_url": pic}

def select_product_with_ai(products, caption=None, img_path=None):
    """ใช้ Gemini เพื่อวิเคราะห์ความเชื่อมโยงของโพสต์กับสินค้าสปอนเซอร์"""
    api_key = os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
    if not api_key:
        print("AI Product Selector: No API key found. Falling back to random selection.")
        return None

    try:
        from google import genai
        from google.genai import types

        client = genai.Client(api_key=api_key, http_options={'timeout': 90.0})
        
        # จัดเตรียมรายการสินค้าที่มีในระบบ
        product_list_str = ""
        for idx, p in enumerate(products):
            product_list_str += f"[{idx}] Name: {p.get('name', '')}, Desc: {p.get('desc', '')}\n"

        prompt = (
            "คุณคือผู้ช่วยระบบ AI Product Selector สำหรับหน้าเพจโซเชียลมีเดียในไทย\n"
            "กรุณาวิเคราะห์ภาพถ่าย และ/หรือ ข้อความแคปชั่นของโพสต์นี้ เพื่อจับคู่เลือกสินค้าสปอนเซอร์ (Affiliate Product) ที่เหมาะสมและเนียนที่สุดจากรายการด้านล่าง:\n"
        )
        if caption:
            prompt += f"ข้อความแคปชั่นโพสต์:\n\"\"\"\n{caption}\n\"\"\"\n\n"

        prompt += (
            "และนี่คือรายการสินค้าสปอนเซอร์ที่มีในระบบ:\n"
            f"{product_list_str}\n"
            "กฎการเลือกสินค้า:\n"
            "- วิเคราะห์โพสต์/รูปภาพเพื่อเลือกสินค้าที่เกี่ยวข้องกันมากที่สุด เช่น โพสต์เรื่องสัตว์เลี้ยงควรคู่กับสินค้าสัตว์เลี้ยง, โพสต์มนุษย์ออฟฟิศ/ปวดหลัง/ทำงานหนักควรคู่กับเบาะรองนั่ง/แผ่นแปะแก้ปวด/แว่นสายตา/ลูทีนบำรุงสายตา, โพสต์อาหารคาวหรือกินเผ็ดร้อนควรคู่กับยาลดกรดหรืออาหารแก้เผ็ด\n"
            "- ให้เลือกดัชนีของสินค้ามาเพียงชิ้นเดียวเท่านั้น\n"
            "- ตอบกลับเป็นดัชนีของสินค้าในเครื่องหมายวงเล็บเหลี่ยมเท่านั้น เช่น [3] (ห้ามเขียนบรรยายความรู้สึกหรือเหตุผล ห้ามมีข้อความอื่นปน)\n"
            "- หากวิเคราะห์แล้วไม่มีสินค้าใดเหมาะสมหรือใกล้เคียงเลย ให้ตอบว่า [None]"
        )

        contents = []
        if img_path and os.path.exists(img_path):
            try:
                with open(img_path, "rb") as f:
                    img_data = f.read()
                ext = os.path.splitext(img_path)[1].lower()
                mime_type = "image/jpeg"
                if ext == ".png":
                    mime_type = "image/png"
                elif ext == ".webp":
                    mime_type = "image/webp"
                elif ext == ".gif":
                    mime_type = "image/gif"
                contents.append(types.Part.from_bytes(data=img_data, mime_type=mime_type))
            except Exception as img_err:
                print(f"AI Product Selector image read error: {img_err}")

        contents.append(prompt)

        # เรียกใช้ Gemini-2.5-flash (เร็วและประหยัด)
        resp = client.models.generate_content(
            model="gemini-flash-latest",
            contents=contents
        )

        result = resp.text.strip()
        print(f"AI Product Selector response: {result}")

        match = re.search(r'\[(\d+)\]', result)
        if match:
            idx = int(match.group(1))
            if 0 <= idx < len(products):
                return products[idx]
        elif "[None]" in result or "None" in result:
            print("AI Product Selector: Decided no product is relevant.")
    except Exception as e:
        print(f"AI Product Selector error: {e}")

    return None

def get_product_comments(caption=None, img_path=None):
    """comments สินค้าหมุนเวียน (ใช้ AI วิเคราะห์และเลือก) — คืนค่าลิสต์ที่มี 1 คอมเมนต์รวมลิงก์ Shopee/Lazada"""
    products, _, _ = _load_excel()
    active = [p for p in products if p["shopee"] and "xxx" not in p["shopee"]]
    if not active:
        return []
    
    # Cap candidates to 25 to prevent token bloat
    if len(active) > 25:
        active = random.sample(active, 25)
        print(f"Capped active products to 25 candidates for AI selector.")
    
    selected_p = None
    if caption or img_path:
        selected_p = select_product_with_ai(active, caption=caption, img_path=img_path)
        
    if not selected_p:
        # หาก AI เลือกไม่ได้ หรือคีย์หมดโควตา — สุ่มเลือกสินค้าที่แอคทีฟมาแนะนำแทน
        print("AI Selector: No relevant product matched. Falling back to random active product.")
        selected_p = random.choice(active)
    else:
        print(f"AI Selector: Selected product -> {selected_p['name']}")

    p = selected_p
    persona = get_persona()
    
    # กำหนดคำลงท้าย
    if persona == "chowchow":
        ending = "ฮะ โฮ่ง!"
    elif persona == "somtam":
        ending = "ค่ะ"
    else:
        ending = "ครับ"

    shopee_url = p.get("shopee", "").strip()
    lazada_url = p.get("lazada", "").strip()
    if "xxx" in shopee_url: shopee_url = ""
    if "xxx" in lazada_url: lazada_url = ""

    if not shopee_url and not lazada_url:
        return []

    # 1. พยายามใช้ AI เขียนข้อความโปรโมทสินค้าแบบธรรมชาติโดยมี caption ช่วยเชื่อมโยง
    ai_comment = generate_comment_with_ai(p, "Shopee & Lazada", persona, caption=caption)
    
    if ai_comment:
        msg = ai_comment
    else:
        # 2. หาก AI ล้มเหลว ให้ใช้ Fallback Template
        price_str = p.get("price", "")
        price_val = f" ราคา {price_str} บาท" if price_str else ""
        templates = [
            f"📍 เผื่อใครถามพิกัดของ {p['name']}{price_val} ที่เห็นในโพสต์นะครับ",
            f"💬 มีคนถามถึง {p['name']}{price_val} บ่อยๆ วางพิกัดไว้ให้ทางนี้เลยครับ",
            f"💡 {p['name']}{price_val} ตัวที่เล่าไป ใครสนใจดูรายละเอียดและสั่งซื้อได้ตรงนี้ครับ",
            f"🛒 ใครหา {p['name']}{price_val} อยู่ แปะลิงก์ร้านค้าไว้ให้เรียบร้อยครับ"
        ]
        msg = random.choice(templates)
        # ปรับสรรพนามและลงท้ายตามเพจ
        if persona == "somtam":
            msg = msg.replace("ครับ", "ค่ะ").replace("นะครับ", "นะคะ")
        elif persona == "chowchow":
            msg = msg.replace("ครับ", "ฮะ โฮ่ง!").replace("นะครับ", "นะฮะ โฮ่ง!")"

    # รวมลิงก์
    links = []
    if shopee_url:
        links.append(f"🧡 Shopee: {shopee_url}")
    if lazada_url:
        links.append(f"💙 Lazada: {lazada_url}")
        
    link_section = "\n".join(links)
    combined_comment = f"{msg}\n\n📌 พิกัดของชิ้นนี้จิ้มได้เลยน้า:\n{link_section}"
    
    return [combined_comment]

def get_website_comment():
    web = _rotate(WEBSITE_VARS)
    if not web:
        return ""
    persona = get_persona()
    if persona == "somtam":
        web = web.replace("ครับ", "ค่ะ")
    elif persona == "chowchow":
        web = web.replace("ครับ", "ฮะ โฮ่ง!")
    return web

def get_all_comments(caption=None, img_path=None):
    """
    คืนค่าคอมเมนต์สูงสุด 2 คอมเมนต์ต่อโพสต์เพื่อไม่ให้สแปม:
    1. คอมเมนต์โปรโมทหลัก (Promo -> Product -> Food)
    2. คอมเมนต์แนะนำเว็บไซต์ (Website shopee-ranking)
    """
    comments = []
    
    # 1. คอมเมนต์หลัก (สูงสุด 1 ชิ้น)
    promo = get_promo_comment()
    if promo:
        comments.append(promo)
    else:
        prod_comments = get_product_comments(caption=caption, img_path=img_path)
        if prod_comments:
            comments.extend(prod_comments)
        else:
            food = get_food_comment()
            if food:
                comments.append(food)
                
    # 2. คอมเมนต์แนะนำเว็บ (สูงสุด 1 ชิ้น)
    web = get_website_comment()
    if web:
        comments.append(web)
        
    return comments

def get_next_scheduled_time(slots):
    """
    Calculate next scheduled publish time in Bangkok timezone (UTC+7).
    Bypassed if IMMEDIATE=true is set in environment.
    """
    if os.environ.get("IMMEDIATE") == "true":
        print("IMMEDIATE environment variable is set to true. Skipping scheduling...")
        return None

    bkk = timezone(timedelta(hours=7))
    now = datetime.now(bkk)
    
    candidates = []
    for day_offset in [0, 1]:
        base_date = now + timedelta(days=day_offset)
        for slot in slots:
            hour, min_val = map(int, slot.split(":"))
            dt = datetime(base_date.year, base_date.month, base_date.day, hour, min_val, tzinfo=bkk)
            # Facebook Graph API requires scheduled_publish_time to be >= 10 mins in future
            if dt > now + timedelta(minutes=10):
                candidates.append(dt)
                
    candidates.sort()
    if candidates:
        target = candidates[0]
        print(f"Calculated next target slot: {target.strftime('%Y-%m-%d %H:%M:%S %Z')} (UNIX: {int(target.timestamp())})")
        return int(target.timestamp())
    return None

